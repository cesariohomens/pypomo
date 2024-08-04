import tkinter as tk
import tkinter.ttk as ttk
import pygame
import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from datetime import datetime
import os
import csv

# Get the directory where the script is located
script_dir = os.path.dirname(os.path.realpath(__file__))

# Change the working directory to the script's directory
os.chdir(script_dir)

# Construct paths to the mp3 and png files
alarm_sound_path = os.path.join(script_dir, "alarm.mp3")
icon_path = os.path.join(script_dir, "pypomo.png")
excel_path = os.path.join(script_dir, "data.xlsx")
config_path = os.path.join(script_dir, "config.csv")

# Initialize pygame mixer
pygame.mixer.init()

# Load the alarm sound
alarm_sound = pygame.mixer.Sound(alarm_sound_path)  # Ensure "alarm.mp3" is in the same directory as this script

# Define global variables for the language configurations
language_list = []
default_language = None

# Load the configuration file and determine the default language
def load_config():
    global language_list, default_language
    with open(config_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            language_list.append(row)
            if row['used'] == '1':
                default_language = row

# Load translations from a CSV file
def load_translations(language_file):
    translations = {}
    language_file_path = os.path.join(script_dir, "translations", language_file)
    with open(language_file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row:
                translations[row[0]] = row[1]
    return translations


# Save language choice to the configuration file
def save_language_choice(selected_language):
    with open(config_path, "w", newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=['language', 'filename', 'used'])
        writer.writeheader()
        for lang in language_list:
            lang['used'] = '1' if lang['filename'] == selected_language['filename'] else '0'
            writer.writerow(lang)

# Update the UI texts with the translations
def update_ui_texts():
    pomodoro_button.config(text=translations['pomodoro'])
    short_break_button.config(text=translations['short-break'])
    long_break_button.config(text=translations['long-break'])
    start_button.config(text=translations['start'])
    pause_button.config(text=translations['pause'] if not paused else translations['resume'])
    stop_button.config(text=translations['stop'])
    reset_button.config(text=translations['reset'])
    minute_up_button.config(text=translations['min-p'])
    minute_down_button.config(text=translations['min-m'])
    second_up_button.config(text=translations['sec-p'])
    second_down_button.config(text=translations['sec-m'])
    project_label.config(text=translations['project'])
    type_label.config(text=translations['type'])
    delete_button_project.config(text=translations['del-project'])
    delete_button_type.config(text=translations['del-type'])
    delete_button_line.config(text=translations['del-line'])
    work_stats_label.config(text=translations['work-stats'])
    break_stats_label.config(text=translations['break-stats'])
    work_day_label.config(text=f"{translations['day']}: {format_duration(work_total_day)}")
    break_day_label.config(text=f"{translations['day']}: {format_duration(break_total_day)}")
    work_checkbox_label.config(text=translations['list-works'])
    break_checkbox_label.config(text=translations['list-breaks'])
    # Update table column headers
    for col in columns:
        table.heading(col, text=translations[col])

# Update the language and save the choice
def update_language(new_language):
    global translations
    translations = load_translations(new_language['filename'])
    update_ui_texts()  # Update UI texts with new translations
    save_language_choice(new_language)  # Save the new language choice to a configuration file

# Initialize the main window
window = tk.Tk()
window.title("PyPomo")

# Load configuration and translations
load_config()
translations = load_translations(default_language['filename'])

# Define global variables for start and end times
var_start = 0
var_end = 0
var_pomodoro = 0

# Function to convert datetime string to datetime object
def parse_datetime(dt_str):
    return dt.datetime.strptime(dt_str, '%d/%m/%Y %I:%M:%S %p') if not pd.isna(dt_str) else None

# Read data from Excel file and write stats
def read_excel_write_stats():
    global project_data, type_data, data_df, work_total_day, work_total_week, work_total_month, work_total_year, break_total_day, break_total_week, break_total_month, break_total_year
    excel_data = pd.ExcelFile(excel_path)
    project_data = excel_data.parse('project')
    type_data = excel_data.parse('type')
    data_df = excel_data.parse('data')

    # Convert 'start_time' and 'end_time' from string to datetime objects
    data_df['start_time'] = data_df['start_time'].apply(parse_datetime)
    data_df['end_time'] = data_df['end_time'].apply(parse_datetime)

    # Calculate 'duration'
    data_df['duration'] = data_df['end_time'] - data_df['start_time']

    # Filter pomodoro rows to get stats
    data_work = data_df[data_df['pomodoro'] == 1]
    data_break = data_df[data_df['pomodoro'] == 0]

    # Calculate total durations for day, week, month, year
    current_date = dt.datetime.now()
    if not data_work.empty:
        work_total_day = data_work[data_work['start_time'].dt.date == current_date.date()]['duration'].sum()
        work_total_week = data_work[data_work['start_time'].dt.isocalendar().week == current_date.isocalendar().week]['duration'].sum()
        work_total_month = data_work[(data_work['start_time'].dt.month == current_date.month) & (data_work['start_time'].dt.year == current_date.year)]['duration'].sum()
        work_total_year = data_work[data_work['start_time'].dt.year == current_date.year]['duration'].sum()
    else:
        work_total_day = work_total_week = work_total_month = work_total_year = pd.Timedelta(0)

    if not data_break.empty:
        break_total_day = data_break[data_break['start_time'].dt.date == current_date.date()]['duration'].sum()
        break_total_week = data_break[data_break['start_time'].dt.isocalendar().week == current_date.isocalendar().week]['duration'].sum()
        break_total_month = data_break[(data_break['start_time'].dt.month == current_date.month) & (data_break['start_time'].dt.year == current_date.year)]['duration'].sum()
        break_total_year = data_break[data_break['start_time'].dt.year == current_date.year]['duration'].sum()
    else:
        break_total_day = break_total_week = break_total_month = break_total_year = pd.Timedelta(0)

# Read Excel and write stats
read_excel_write_stats()

# Function to refresh the table based on checkbox states
def refresh_table():
    filtered_data = data_df.copy()
    filtered_data = filtered_data.sort_values(by='end_time', ascending=False)
    if list_works_var.get() and list_breaks_var.get():
        pass
    elif list_works_var.get():
        filtered_data = filtered_data[filtered_data['pomodoro'] == 1]
    elif list_breaks_var.get():
        filtered_data = filtered_data[filtered_data['pomodoro'] == 0]
    else:
        filtered_data = filtered_data[0:0]
    for item in table.get_children():
        table.delete(item)
    for idx, row in filtered_data.iterrows():
        start_time = row['start_time'].strftime("%d/%m/%Y %I:%M:%S %p") if row['start_time'] else ''
        end_time = row['end_time'].strftime("%d/%m/%Y %I:%M:%S %p") if row['end_time'] else ''
        duration = format_duration(row['duration'])
        project = project_data.loc[project_data['id'] == row['project_id'], 'description'].iloc[0] if row['project_id'] else ''
        type_val = type_data.loc[type_data['id'] == row['type_id'], 'type'].iloc[0] if row['type_id'] else ''
        pomodoro = "Work" if row['pomodoro'] == 1 else "Break"
        table.insert("", tk.END, values=(start_time, end_time, duration, project, type_val, pomodoro))

# Format duration as hh:mm:ss
def format_duration(td):
    return f"{int(td.total_seconds() // 3600):02d}:{int((td.total_seconds() % 3600) // 60):02d}:{int(td.total_seconds() % 60):02d}"

# Create lists for dropdown menus
project_list = project_data['description'].tolist()
type_list = type_data['type'].tolist()

# Maps to store id corresponding to the dropdown selections
project_id_map = dict(zip(project_data['description'], project_data['id']))
type_id_map = dict(zip(type_data['type'], type_data['id']))

# Initialize timer variables
timer = None  # Timer control variable
default_time = {"Pomodoro": 25*60, "Short Break": 5*60, "Long Break": 15*60}
current_mode = "Pomodoro"
minutes = 25
seconds = 0
running = False
paused = False

# Function to update the timer label
def update_label():
    label.config(text=f"{minutes:02d}:{seconds:02d}")

# Function to change timer mode and reset time
def change_mode(mode):
    global current_mode, running, paused, minutes, seconds
    current_mode = mode
    minutes, seconds = divmod(default_time[current_mode], 60)
    update_label()
    window.configure(bg=mode_colors[mode])
    label.configure(bg=mode_colors[mode])
    dropdown_frame.configure(bg=mode_colors[mode])
    project_label.configure(bg=mode_colors[mode])
    type_label.configure(bg=mode_colors[mode])
    delete_frame.configure(bg=mode_colors[mode])
    delete_button_project.configure(bg=mode_colors[mode])
    delete_button_type.configure(bg=mode_colors[mode])
    delete_button_line.configure(bg=mode_colors[mode])
    time_work_frame.configure(bg=mode_colors[mode])
    work_stats_label.configure(bg=mode_colors[mode])
    work_day_label.configure(bg=mode_colors[mode])
    work_week_label.configure(bg=mode_colors[mode])
    work_month_label.configure(bg=mode_colors[mode])
    work_year_label.configure(bg=mode_colors[mode])
    work_list_checkbox.configure(bg=mode_colors[mode])
    work_checkbox_label.configure(bg=mode_colors[mode])
    time_break_frame.configure(bg=mode_colors[mode])
    break_stats_label.configure(bg=mode_colors[mode])
    break_day_label.configure(bg=mode_colors[mode])
    break_week_label.configure(bg=mode_colors[mode])
    break_month_label.configure(bg=mode_colors[mode])
    break_year_label.configure(bg=mode_colors[mode])
    break_list_checkbox.configure(bg=mode_colors[mode])
    break_checkbox_label.configure(bg=mode_colors[mode])
    if running or paused:
        stop_timer()

# Function to increment/decrement minutes
def adjust_minutes(m):
    global minutes
    minutes += m
    if minutes < 0:
        minutes = 59
    if minutes > 59:
        minutes = 0
    update_label()

# Function to increment/decrement seconds
def adjust_seconds(s):
    global seconds
    seconds += s
    if seconds < 0:
        seconds = 59
    elif seconds > 59:
        seconds = 0
    update_label()

# Timer countdown function
def countdown():
    global minutes, seconds, running, paused
    if running and not paused:
        if seconds > 0:
            seconds -= 1
        elif minutes > 0:
            minutes -= 1
            seconds = 59
        else:
            stop_timer()
            alarm_sound.play()
            return
        label.config(text=f"{minutes:02d}:{seconds:02d}")
        window.after(1000, countdown)

def show_custom_warning(message_key):
    message = translations.get(message_key, "Warning")  # Use the key to get the translated message
    title = translations.get("warning-title", "Warning")
    popup = tk.Toplevel(window)
    popup.title(title)
    popup.configure(bg=mode_colors[current_mode])
    popup.grab_set()
    popup_width = 300
    popup_height = 100
    center_x = int(window.winfo_x() + window.winfo_width() / 2 - popup_width / 2)
    center_y = int(window.winfo_y() + window.winfo_height() / 2 - popup_height / 2)
    popup.geometry(f"{popup_width}x{popup_height}+{center_x}+{center_y}")
    label = tk.Label(popup, text=message, fg="white", bg=mode_colors[current_mode], font=("Arial", 12))
    label.pack(pady=(10, 5), padx=10)
    dismiss_button = tk.Button(popup, text=translations.get("ok", "OK"), command=lambda: [popup.grab_release(), popup.destroy()], fg="black", bg="white", relief="raised")
    dismiss_button.pack(pady=(0, 10))


# Add Project and Types to Excel
def add_to_excel_sheet(file_path, sheet_name, new_data):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    next_row = sheet.max_row + 1
    for col, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col, value=value)
    workbook.save(file_path)

# Update dropdowns after inserting Project and Types to Excel
def update_dropdowns():
    global project_data, type_data, project_list, type_list
    excel_data = pd.ExcelFile(excel_path)
    project_data = excel_data.parse('project')
    type_data = excel_data.parse('type')
    project_list = project_data['description'].tolist()
    type_list = type_data['type'].tolist()
    project_menu['values'] = project_list
    type_menu['values'] = type_list

# Function for Start button
def start_timer():
    global running, paused, var_start, var_pomodoro, project_id_map, type_id_map
    project_name = project_menu.get()
    type_name = type_menu.get()
    if project_name and project_name not in project_list:
        new_project_id = project_data['id'].max() + 1 if not project_data.empty else 1
        add_to_excel_sheet(excel_path, 'project', [new_project_id, project_name])
    if type_name and type_name not in type_list:
        new_type_id = type_data['id'].max() + 1 if not type_data.empty else 1
        add_to_excel_sheet(excel_path, 'type', [new_type_id, type_name])
    update_dropdowns()
    project_id_map = dict(zip(project_data['description'], project_data['id']))
    type_id_map = dict(zip(type_data['type'], type_data['id']))
    if not project_name or not type_name:
        show_custom_warning("select-project-type")
        return
    var_start = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
    var_pomodoro = 1 if current_mode == "Pomodoro" else 0
    if not running or paused:
        running = True
        paused = False
        countdown()

# Function to format datetime object to Excel date
def to_excel_datetime(dt_str):
    return datetime.strptime(dt_str, "%d/%m/%Y %I:%M:%S %p")

# Function to append data to an existing Excel sheet
def append_to_excel(file_path, sheet_name, data_dict):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    next_row = sheet.max_row + 1
    date_format = "DD/MM/YYYY HH:MM:SS AM/PM"
    for col, (key, value) in enumerate(data_dict.items(), start=1):
        cell = sheet.cell(row=next_row, column=col)
        if key in ['start_time', 'end_time']:
            cell.value = to_excel_datetime(value)
            cell.number_format = date_format
        else:
            cell.value = value
    workbook.save(file_path)

# Function for Stop button
def stop_timer():
    global running, paused, var_start, var_end, var_pomodoro, data_df
    if not var_start:
        return
    var_end = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
    running = False
    paused = False
    change_mode(current_mode)
    new_data = {
        1: var_start,
        2: var_end,
        3: project_id_map[project_menu.get()],
        4: type_id_map[type_menu.get()],
        5: var_pomodoro
    }
    append_to_excel(excel_path, 'data', new_data)
    read_excel_write_stats()
    work_day_label.config(text=f"{translations['day']}: {format_duration(work_total_day)}")
    work_week_label.config(text=f"{translations['week']}: {format_duration(work_total_week)}")
    work_month_label.config(text=f"{translations['month']}: {format_duration(work_total_month)}")
    work_year_label.config(text=f"{translations['year']}: {format_duration(work_total_year)}")
    break_day_label.config(text=f"{translations['day']}: {format_duration(break_total_day)}")
    break_week_label.config(text=f"{translations['week']}: {format_duration(break_total_week)}")
    break_month_label.config(text=f"{translations['month']}: {format_duration(break_total_month)}")
    break_year_label.config(text=f"{translations['year']}: {format_duration(break_total_year)}")
    data_df.sort_values(by='end_time', ascending=False, inplace=True)
    for item in table.get_children():
        table.delete(item)
    refresh_table()
    var_start = 0
    var_end = 0

# Pause the timer
def pause_timer():
    global paused, var_start, var_end
    if not paused:
        paused = True
        var_end = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
        append_to_excel(excel_path, 'data', {
            1: var_start,
            2: var_end,
            3: project_id_map[project_menu.get()],
            4: type_id_map[type_menu.get()],
            5: 1
        })
        pause_button.config(text=translations['resume'])
        start_break_timer()
    else:
        resume_timer()

def start_break_timer():
    global var_start
    var_start = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")

# Resume timer
def resume_timer():
    global var_start, var_end, paused
    var_end = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
    append_to_excel(excel_path, 'data', {
        1: var_start,
        2: var_end,
        3: project_id_map[project_menu.get()],
        4: type_id_map[type_menu.get()],
        5: 0
    })
    var_start = dt.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
    paused = False
    countdown()
    pause_button.config(text=translations['pause'])

# Function for Reset button
def reset_timer():
    global running, paused, var_start, var_end
    running = False
    paused = False
    var_start = 0
    var_end = 0
    change_mode("Pomodoro")

# Function for Delete Project button
def delete_project():
    selected_project = project_menu.get()
    if not selected_project:
        show_custom_warning("no-project-selected")
        return
    project_id = project_id_map.get(selected_project)
    if project_id is None:
        show_custom_warning("no-project-found")
        return
    if any(data_df['project_id'] == project_id):
        show_custom_warning("cannot-remove-project")
        return
    global project_data
    project_data = project_data[project_data['id'] != project_id]
    workbook = load_workbook(excel_path)
    sheet = workbook['project']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == project_id:
            sheet.delete_rows(row[0].row)
            break
    workbook.save(excel_path)
    read_excel_write_stats()
    project_list = project_data['description'].tolist()
    project_menu['values'] = project_list
    project_menu.set('')
    show_custom_warning("project-removed")

# Function for Delete Type button
def delete_type():
    selected_type = type_menu.get()
    if not selected_type:
        show_custom_warning("no-type-selected")
        return
    type_id = type_id_map.get(selected_type)
    if type_id is None:
        show_custom_warning("no-type-found")
        return
    if any(data_df['type_id'] == type_id):
        show_custom_warning("cannot-remove-type")
        return
    global type_data
    type_data = type_data[type_data['id'] != type_id]
    workbook = load_workbook(excel_path)
    sheet = workbook['type']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == type_id:
            sheet.delete_rows(row[0].row)
            break
    workbook.save(excel_path)
    read_excel_write_stats()
    type_list = type_data['type'].tolist()
    type_menu['values'] = type_list
    type_menu.set('')
    show_custom_warning("type-removed")

def delete_line():
    selected_item = table.selection()
    if not selected_item:
        show_custom_warning("select-line")
        return
    selected_row = table.item(selected_item)['values']
    if not selected_row:
        show_custom_warning("invalid-selection")
        return
    selected_start_time = selected_row[0]
    workbook = load_workbook(excel_path)
    sheet = workbook['data']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and dt.datetime.strptime(row[0].value, "%d/%m/%Y %I:%M:%S %p").strftime("%d/%m/%Y %I:%M:%S %p") == selected_start_time:
            sheet.delete_rows(row[0].row)
            workbook.save(excel_path)
            break
    read_excel_write_stats()
    work_day_label.config(text=f"{translations['day']}: {format_duration(work_total_day)}")
    work_week_label.config(text=f"{translations['week']}: {format_duration(work_total_week)}")
    work_month_label.config(text=f"{translations['month']}: {format_duration(work_total_month)}")
    work_year_label.config(text=f"{translations['year']}: {format_duration(work_total_year)}")
    break_day_label.config(text=f"{translations['day']}: {format_duration(break_total_day)}")
    break_week_label.config(text=f"{translations['week']}: {format_duration(break_total_week)}")
    break_month_label.config(text=f"{translations['month']}: {format_duration(break_total_month)}")
    break_year_label.config(text=f"{translations['year']}: {format_duration(break_total_year)}")
    refresh_table()
    show_custom_warning("line-removed")

# Define mode colors
mode_colors = {
    "Pomodoro": "#ba4949",
    "Short Break": "#38858a",
    "Long Break": "#397097"
}

# Set the application icon
window.iconphoto(True, tk.PhotoImage(file=icon_path))

# Get the screen width and height of the primary display
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()

# Define window dimensions
window_width = 1000
window_height = 600

# Calculate x and y coordinates for the Tk root window
center_x = int(screen_width / 2 - window_width / 2)
center_y = int(screen_height / 2 - window_height / 2)

# Set the position of the window to the center of the screen
window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

# Styling
window.configure(bg=mode_colors[current_mode])

# Timer label
label = tk.Label(window, text="25:00", font=("Arial", 48), fg="white", bg=mode_colors[current_mode])
label.pack(pady=(20, 10))

# Frame for mode buttons and control buttons
buttons_frame = tk.Frame(window, bg=mode_colors[current_mode])
buttons_frame.pack(pady=(0, 10))

# Mode buttons
pomodoro_button = tk.Button(buttons_frame, text=translations['pomodoro'], command=lambda: change_mode("Pomodoro"), fg="white", bg="#ba4949", relief="flat")
short_break_button = tk.Button(buttons_frame, text=translations['short-break'], command=lambda: change_mode("Short Break"), fg="white", bg="#38858a", relief="flat")
long_break_button = tk.Button(buttons_frame, text=translations['long-break'], command=lambda: change_mode("Long Break"), fg="white", bg="#397097", relief="flat")

pomodoro_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
short_break_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
long_break_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Control buttons
start_button = tk.Button(buttons_frame, text=translations['start'], command=start_timer, fg="black", bg="white", relief="raised")
pause_button = tk.Button(buttons_frame, text=translations['pause'], command=pause_timer, fg="black", bg="white", relief="raised")
stop_button = tk.Button(buttons_frame, text=translations['stop'], command=stop_timer, fg="black", bg="white", relief="raised")
reset_button = tk.Button(buttons_frame, text=translations['reset'], command=reset_timer, fg="black", bg="white", relief="raised")

start_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
pause_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
stop_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
reset_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Minute and second adjustment buttons frame
adjustment_frame = tk.Frame(window, bg=mode_colors[current_mode])
adjustment_frame.pack(pady=(0, 10))

# Minute adjustment buttons
minute_up_button = tk.Button(adjustment_frame, text=translations['min-p'], command=lambda: adjust_minutes(1), fg="black", bg="white", relief="raised")
minute_down_button = tk.Button(adjustment_frame, text=translations['min-m'], command=lambda: adjust_minutes(-1), fg="black", bg="white", relief="raised")
minute_up_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
minute_down_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Second adjustment buttons
second_up_button = tk.Button(adjustment_frame, text=translations['sec-p'], command=lambda: adjust_seconds(1), fg="black", bg="white", relief="raised")
second_down_button = tk.Button(adjustment_frame, text=translations['sec-m'], command=lambda: adjust_seconds(-1), fg="black", bg="white", relief="raised")
second_up_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
second_down_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Dropdown menus frame
dropdown_frame = tk.Frame(window, bg=mode_colors[current_mode])
dropdown_frame.pack(pady=(0, 10))

# Project Dropdown
project_label = tk.Label(dropdown_frame, text=translations['project'], bg=mode_colors[current_mode], fg="white")
project_label.pack(side=tk.LEFT)
project_menu = ttk.Combobox(dropdown_frame, values=project_list)
project_menu.pack(side=tk.LEFT, padx=(5, 20))

# Type Dropdown
type_label = tk.Label(dropdown_frame, text=translations['type'], bg=mode_colors[current_mode], fg="white")
type_label.pack(side=tk.LEFT)
type_menu = ttk.Combobox(dropdown_frame, values=type_list)
type_menu.pack(side=tk.LEFT)

# Language selection frame
language_frame = tk.Frame(window, bg=mode_colors[current_mode])
language_frame.place(relx=1.0, x=-20, y=20, anchor='ne')  # Adjust placement as needed

# Default language selection based on config
language_var = tk.StringVar(value=default_language['language'])
language_menu = ttk.Combobox(language_frame, textvariable=language_var, values=[lang['language'] for lang in language_list], state="readonly")
language_menu.pack(side=tk.LEFT)
language_menu.bind("<<ComboboxSelected>>", lambda event: update_language(next((lang for lang in language_list if lang['language'] == language_var.get()), None)))

# Frame for mode buttons and control buttons
delete_frame = tk.Frame(window, bg=mode_colors[current_mode])
delete_frame.pack(pady=(0, 10))

# Control buttons
delete_button_project = tk.Button(delete_frame, text=translations['del-project'], command=delete_project, fg="white", bg=mode_colors[current_mode], relief="flat")
delete_button_type = tk.Button(delete_frame, text=translations['del-type'], command=delete_type, fg="white", bg=mode_colors[current_mode], relief="flat")
delete_button_line = tk.Button(delete_frame, text=translations['del-line'], command=delete_line, fg="white", bg=mode_colors[current_mode], relief="flat")

delete_button_project.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
delete_button_type.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
delete_button_line.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Time work labels frame
time_work_frame = tk.Frame(window, bg=mode_colors[current_mode])
time_work_frame.pack(pady=(0, 10))

# Define a BooleanVar for the checkbox list works
list_works_var = tk.BooleanVar(value=True)  # Default value as checked

# Define a BooleanVar for the checkbox list breaks
list_breaks_var = tk.BooleanVar(value=False)  # Default value as checked

# Time work labels
work_stats_label = tk.Label(time_work_frame, text=translations['work-stats'], font=("Arial", 12, "bold"), bg=mode_colors[current_mode], fg="white")
work_day_label = tk.Label(time_work_frame, text=f"{translations['day']}: {format_duration(work_total_day)}", bg=mode_colors[current_mode], fg="white")
work_week_label = tk.Label(time_work_frame, text=f"{translations['week']}: {format_duration(work_total_week)}", bg=mode_colors[current_mode], fg="white")
work_month_label = tk.Label(time_work_frame, text=f"{translations['month']}: {format_duration(work_total_month)}", bg=mode_colors[current_mode], fg="white")
work_year_label = tk.Label(time_work_frame, text=f"{translations['year']}: {format_duration(work_total_year)}", bg=mode_colors[current_mode], fg="white")
work_list_checkbox = tk.Checkbutton(time_work_frame, var=list_works_var, command=refresh_table, bg=mode_colors[current_mode])
work_checkbox_label = tk.Label(time_work_frame, text=translations['list-works'], bg=mode_colors[current_mode], fg="white", anchor='w')

work_stats_label.pack(side=tk.LEFT, padx=(10, 10))
work_day_label.pack(side=tk.LEFT, padx=(10, 10))
work_week_label.pack(side=tk.LEFT, padx=(10, 10))
work_month_label.pack(side=tk.LEFT, padx=(10, 10))
work_year_label.pack(side=tk.LEFT, padx=(10, 10))
work_list_checkbox.pack(side=tk.LEFT, padx=(10, 10))
work_checkbox_label.pack(side=tk.LEFT, padx=(10, 10))

# Time break labels frame
time_break_frame = tk.Frame(window, bg=mode_colors[current_mode])
time_break_frame.pack(pady=(0, 10))

# Time break labels
break_stats_label = tk.Label(time_break_frame, text=translations['break-stats'], font=("Arial", 12, "bold"), bg=mode_colors[current_mode], fg="white")
break_day_label = tk.Label(time_break_frame, text=f"{translations['day']}: {format_duration(break_total_day)}", bg=mode_colors[current_mode], fg="white")
break_week_label = tk.Label(time_break_frame, text=f"{translations['week']}: {format_duration(break_total_week)}", bg=mode_colors[current_mode], fg="white")
break_month_label = tk.Label(time_break_frame, text=f"{translations['month']}: {format_duration(break_total_month)}", bg=mode_colors[current_mode], fg="white")
break_year_label = tk.Label(time_break_frame, text=f"{translations['year']}: {format_duration(break_total_year)}", bg=mode_colors[current_mode], fg="white")
break_list_checkbox = tk.Checkbutton(time_break_frame, var=list_breaks_var, command=refresh_table, bg=mode_colors[current_mode])
break_checkbox_label = tk.Label(time_break_frame, text=translations['list-breaks'], bg=mode_colors[current_mode], fg="white", anchor='w')

break_stats_label.pack(side=tk.LEFT, padx=(10, 10))
break_day_label.pack(side=tk.LEFT, padx=(10, 10))
break_week_label.pack(side=tk.LEFT, padx=(10, 10))
break_month_label.pack(side=tk.LEFT, padx=(10, 10))
break_year_label.pack(side=tk.LEFT, padx=(10, 10))
break_list_checkbox.pack(side=tk.LEFT, padx=(10, 10))
break_checkbox_label.pack(side=tk.LEFT, padx=(10, 10))

# Add a frame for the table
table_frame = tk.Frame(window)
table_frame.pack(fill='both', expand=True)

# Create the table
columns = ("start-time", "end-time", "duration", "project", "type", "pomodoro")  # Use translation keys
table = ttk.Treeview(table_frame, columns=columns, show='headings')

# Define the column headings and widths
column_widths = [150, 150, 100, 100, 100, 100]  # Adjust these values as needed
for col, width in zip(columns, column_widths):
    table.heading(col, text=translations[col])  # Use the exact keys with hyphens
    table.column(col, width=width)

# Sort the DataFrame by 'end_time' in descending order
data_df = data_df.sort_values(by='end_time', ascending=False)

# Insert sorted data into the table
refresh_table()

# Add a scrollbar
scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

table.pack(fill='both', expand=True)

# Modify button command assignments
start_button.config(command=start_timer)
stop_button.config(command=stop_timer)
reset_button.config(command=reset_timer)

# Run the application
window.mainloop()